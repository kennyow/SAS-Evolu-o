import openpyxl

# Load the Excel file
file_path = "SAS_EVOLUCAO_2023.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Create a dictionary to store data for different values
data_dict = {}

# Iterate through the rows in the E column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
    cell_value = row[0].value
    if cell_value not in data_dict:
        data_dict[cell_value] = []
    data_dict[cell_value].append([cell.value for cell in sheet[row[0].row]])

# Create a new workbook
new_workbook = openpyxl.Workbook()
new_workbook.remove(new_workbook.active)  # Remove default sheet

# Add sheets for each unique value
for value, entries in data_dict.items():
    new_sheet = new_workbook.create_sheet(title=value)
    for entry in entries:
        new_sheet.append(entry)

# Save the new workbook with all sheets
new_file_path = "TURMAS.xlsx"
new_workbook.save(new_file_path)
new_workbook.close()

# Close the original workbook
workbook.close()
