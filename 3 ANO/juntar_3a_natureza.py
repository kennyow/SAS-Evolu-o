import os
import xlwings as xw
import openpyxl
import subprocess
from time import sleep


#⁡⁢⁢⁣​‌‍‌⁡⁣⁢⁣𝕀𝕄ℙ𝕃𝔼𝕄𝔼ℕ𝕋𝕀ℕ𝔾 𝕋ℍ𝔼 𝔻𝕀ℂ𝕋𝕀𝕆ℕ𝔸ℝ𝕐⁡​⁡
###################################################################################################

# Excel path
arquivo_excel = "SAS_EVOLUCAO_2023.xlsx"

# Load the Excel file
wb1 = openpyxl.load_workbook(arquivo_excel)
sheet2 = wb1.active  # Active sheet

# Save all data for each line
dados_main = {}

# # Iterate through all lines
for linha_numero, column in enumerate(sheet2.iter_rows(min_row=2, values_only=True), start=2):
    email_column_b = column[2]  # Column ⁡⁣⁣⁢C⁡
    matricula_column_d = column[0]  # Column ⁡⁣⁣⁢A⁡

    # Store the values in the dictionary
    dados_main[linha_numero] = [email_column_b, matricula_column_d]

# Create a dictionary to store the values from the column A 
result_dict = {}

#⁡⁣⁣⁢STUDENT MANUAL EXAM.:Iterate throught the cells in column F to determine⁡
for row in sheet2.iter_rows(min_row=2, min_col=6, max_col=6):
    for cell in row:
        if cell.value == 'S':
            # If the value is 'S', retrieve the value from column A in the same row
            col_a_value = sheet2.cell(row=cell.row, column=4).value
            # Add to the dictionary as the key
            result_dict[cell.row] = col_a_value

# Show the dictionary with the results
print(result_dict.values())

# Closing the excel file
wb1.close()


#⁡⁣⁢⁣​‌‍‌𝕆ℝ𝔻𝔼ℕ𝔸𝕋𝕀ℕ𝔾 𝕋ℍ𝔼 𝔻𝕆𝕎ℕ𝕃𝕆𝔸𝔻𝕊 𝕋𝕆 𝔾𝔼𝕋 𝕋ℍ𝔼 𝔼𝕏ℂ𝔼𝕃 𝕃𝔸𝕊𝕋 𝔽𝕀𝕃𝔼𝕊​⁡
#############################################################################################

# Directory where you download files
download_directory = r"C:\Users\citee\Downloads"

# List files in the download directory
files = os.listdir(download_directory)

# Filter out directories and keep only files
files = [f for f in files if os.path.isfile(os.path.join(download_directory, f))]

# Sort files by modification time in ⁡⁣⁣⁢d͟e͟s͟c͟e͟n͟d͟i͟n͟g o͟r͟d͟e͟r⁡ (most recent first)
files.sort(key=lambda x: os.path.getmtime(os.path.join(download_directory, x)), reverse=True)

# Check if there are any downloaded files
if len(files) > 0:
    # Get the name of the ⁡⁣⁣⁢m͟o͟s͟t r͟e͟c͟e͟n͟t͟l͟y d͟o͟w͟n͟l͟o͟a͟d͟e͟d f͟i͟l͟e⁡
    last_downloaded_file2 = files[0]
    input_string = str(last_downloaded_file2[:-4])

    # Print and save the name of the last downloaded file
    print("Last Downloaded File:", input_string)

else:
    print("No downloaded files found in the directory.")

# ⁡⁣⁣⁢Spli⁡⁣⁣⁢t the string by underscore "_"⁡
split_result = input_string.split("_")

pros= str(split_result[4])
if len(pros) > 2:
    pros = str(split_result[5])
avaliacao = str(split_result[1]).upper()
materia = str(split_result[3])
ano = str(pros[:-1])

# Print the result
print("Split Result:", split_result)
print("Year:", ano)
print("Avaliação:", avaliacao)
print(len(avaliacao))

# ⁡⁢⁣⁣OPENING THE SPECIFIC FOLDER WITH CLASSES GROUPS⁡

# Construct the path using the year
original_path = r'G:\Drives compartilhados\_Anos Finais\Anos Finais - 2023\Notas de Avaliações\XXº Ano\II TRIMESTRE\XXº ano - WW - II TRIMESTRE'

# Replace the values
new_path2 = original_path.replace('XX', ano)
new_path = new_path2.replace('WW', avaliacao)

print(new_path)


#⁡⁣⁢⁣​‌‍‌​‌‌‍𝕌𝕊𝔼ℝ'𝕊 𝔾𝕌𝕀𝔻𝔼 𝕋𝕆 ℂℍ𝕆𝕆𝕊𝔼 ℍ𝕆𝕎 𝕄𝔸ℕ𝕐 𝕎𝕆ℝ𝕂𝕊ℍ𝔼𝔼𝕋𝕊​⁡
#############################################################################################
download_directory = r"C:\Users\citee\Downloads"

# List all files in the download directory with their timestamps
files = [(filename, os.path.getmtime(os.path.join(download_directory, filename))) for filename in os.listdir(download_directory)]

# Sort the files by their timestamps in descending order (most recent first)
files.sort(key=lambda x: x[1], reverse=True)

# Create a new workbook to store the extracted data
combined_workbook = openpyxl.Workbook()
combined_sheet = combined_workbook.active

# Initialize a flag to indicate if it's the first file being processed
first_file = True

# Empty dictionary to allocate the subjects
Matérias = []
numb = int(input("Quantidade de Planilhas para compactar: \n"))

# Process the downloaded files
for i, (filename, _) in enumerate(files[:numb]):
    if filename.endswith(".xls"):
        file_path = os.path.join(download_directory, filename)
        
        # Open the Excel file using xlwings
        wb = xw.Book(file_path)
        
        # Allow editing in Excel
        wb.api.Interactive = True
        
        # Extract data from the active sheet
        sheet = wb.sheets.active
        data = sheet.used_range.value
        
        # If it's not the first file, skip the header row
        if not first_file:
            data = data[1:]
        else:
            first_file = False
        
        # Append data to the combined sheet
        for row in data:
            # Add an empty column "A"
            '''row = [''] + list(row)'''
            combined_sheet.append(row)

             # Add distinct value from column "D" to Matérias list
            if row[3] not in Matérias:
                Matérias.append(row[3])
        
        # Close the workbook
        wb.close()

        print(f"Processed data from '{filename}'")

# Save the combined data to a new Excel file
output_path = os.path.join(download_directory, f"Turmas_Unidas_{ano}_{materia}.xlsx")
combined_workbook.save(output_path)
combined_workbook.close()

# ​‌‌‍⁡⁣⁢⁣𝔼𝔻𝕀𝕋𝕀ℕ𝔾 𝕋ℍ𝔼 𝕆𝕌𝕋ℙ𝕌𝕋_ℙ𝔸𝕋ℍ​⁡
##########################################################################################

# Open the saved workbook to adjust column width
saved_workbook = openpyxl.load_workbook(output_path)
saved_sheet = saved_workbook.active

# Iterate through all the cells in column C and delete them
for row in saved_sheet.iter_rows(min_row=1, min_col=3, max_col=3):
    for cell in row:
        cell.value = None



# Adjust the width of column "B" to match the widest content
max_length = max([len(str(cell.value)) for row in saved_sheet for cell in row])
saved_sheet.column_dimensions['B'].width = max_length

# Replace "--" with ⁡⁣⁣⁢"REPOSIÇÃO"⁡ in column "H"
for row in saved_sheet.iter_rows(min_row=2, min_col=8, max_col=8):
    for cell in row:
        if cell.value == "--":
            cell.value = "REPOSIÇÃO"

# Looking for the students that do manual tests
for cell in saved_sheet['C']:
    col_a_value = cell.value
    if col_a_value in result_dict.values():
        # If the value is in the dictionary, write 'MANUAL' in column H on the same row.
        saved_sheet.cell(row=cell.row, column=8, value='MANUAL')

# ⁡⁣⁣⁢Columns to delete (K, L, M, N)⁡
colunas_para_excluir = ['K', 'L', 'M', 'N']

for coluna in colunas_para_excluir:
    for linha in saved_sheet.iter_rows(min_row=2, min_col=11, max_col=20):
        for cell in linha:
            cell.value = None      

#Delete the first value:
Matérias.pop(0)

# ⁡⁣⁢⁣​‌‌‍𝕊𝔼ℙ𝔸ℝ𝔸𝕋𝕀ℕ𝔾 𝕋ℍ𝔼 𝕊ℍ𝔼𝔼𝕋𝕊​⁡
#####################################################################################

# Create separate sheets for each matéria and transfer rows accordingly
for matéria in Matérias:
    if matéria == 'Matemática':
        new_sheet = saved_workbook.create_sheet(title='Matemática')
        
        for row in saved_sheet.iter_rows(min_row=1, values_only=True):
            if row[3] == matéria:
                new_sheet.append(row)
                # Itere pelas células na coluna H e copie os valores para a coluna K
                for row in range(1, new_sheet.max_row + 1):
                    valor_h = new_sheet.cell(row=row, column=8).value
                    new_sheet.cell(row=row, column=11, value=valor_h) 


                # Verifique se o valor na coluna K é igual a 'REPOSIÇÃO' e substitua por 0
                for row in range(2, new_sheet.max_row + 1):
                    valor_k = new_sheet.cell(row=row, column=11).value  # Valor na coluna K
                    if valor_k == 'REPOSIÇÃO':
                        new_sheet.cell(row=row, column=11, value=0)

# Inicialize uma variável para armazenar o maior valor encontrado
maior_valor = None

# Percorra a coluna K e encontre o maior valor
for row in new_sheet.iter_rows(min_row=2, min_col=11, max_col=11):
    for cell in row:
        if cell.value is not None:
            if maior_valor is None or cell.value > maior_valor:
                maior_valor = cell.value
                
# Percorra a coluna K e divida todos os valores pelo maior valor 
for row in new_sheet.iter_rows(min_row=1, min_col=11, max_col=11):
    for cell in row:
        if cell.value is not None:
            cell.value = round(float((cell.value * 10) / maior_valor), 2)
            
            
            

print(f'\033[91m"MAIOR VALOR ENCONTRADO EM MATEMÁTICA:"\033[0m {maior_valor}')                

###################################################################################
new_sheet = saved_workbook.create_sheet(title='Natureza')


# Inicialize uma variável para rastrear o maior valor na coluna H
maior_valor_h = 0
for row in saved_sheet.iter_rows(min_row=2, values_only=True):
        if row[3] != 'Matemática':
            new_sheet.append(row)

     
# Itere pelas células na coluna H em grupos de 3
for row in range(1, saved_sheet.max_row, 3):
    cell1 = new_sheet.cell(row=row, column=8)  # Primeira célula da soma na coluna H
    cell2 = new_sheet.cell(row=row + 1, column=8)  # Segunda célula da soma
    cell3 = new_sheet.cell(row=row + 2, column=8)  # Terceira célula da soma
    
    # Verifique se o valor da célula é 'REPOSIÇÃO' e substitua por 0
    if cell1.value == 'REPOSIÇÃO':
        cell1.value = 0
    if cell2.value == 'REPOSIÇÃO':
        cell2.value = 0
    if cell3.value == 'REPOSIÇÃO':
        cell3.value = 0
    
    # Converta os valores da célula em inteiros, considerando células vazias como 0
    value1 = int(cell1.value) if cell1.value is not None else 0
    value2 = int(cell2.value) if cell2.value is not None else 0
    value3 = int(cell3.value) if cell3.value is not None else 0
    
    soma = value1 + value2 + value3
  
    # Salve o resultado da soma na célula correspondente na coluna K
    new_sheet.cell(row=row, column=11, value=soma)
    new_sheet.cell(row=row + 1, column=11, value=soma)
    new_sheet.cell(row=row + 2, column=11, value=soma)


# Inicialize uma variável para armazenar o maior valor encontrado
maior_valor = None

# Percorra a coluna K e encontre o maior valor
for row in new_sheet.iter_rows(min_row=2, min_col=11, max_col=11):
    for cell in row:
        if cell.value is not None:
            if maior_valor is None or cell.value > maior_valor:
                maior_valor = cell.value

print(f'\033[91m"MAIOR VALOR ENCONTRADO EM NATUREZA:"\033[0m {maior_valor}')   

# Percorra a coluna K e divida todos os valores pelo maior valor 
for row in new_sheet.iter_rows(min_row=1, min_col=11, max_col=11):
    for cell in row:
        if cell.value is not None:
            cell.value = round(float((cell.value * 10) / maior_valor), 2)

##################################################################################

# Crie uma nova planilha para armazenar os dados combinados
new_sheet2 = saved_workbook.create_sheet(title='TOTAL')


# Selecione as duas abas que você deseja combinar
aba1 = saved_workbook['Matemática']
aba2 = saved_workbook['Natureza']

# Copie os dados da primeira aba para a nova planilha
for linha in aba1.iter_rows():
    nova_linha = [c.value for c in linha]
    new_sheet2.append(nova_linha)

# Copie os dados da segunda aba para a nova planilha
for linha in aba2.iter_rows():
    nova_linha = [c.value for c in linha]
    new_sheet2.append(nova_linha)




# Remove the default sheet created by openpyxl
default_sheet = saved_workbook.get_sheet_by_name('Sheet')
saved_workbook.remove_sheet(default_sheet)

saved_workbook.save(output_path)
saved_workbook.close()

print(f"Combined data saved to '{output_path}'.")
print("Distinct values from column D:")
print(Matérias)

#OPEN PATH


path = r"G:\Drives compartilhados\_Anos Finais\Anos Finais - 2023\Notas de Avaliações"

try:
    os.startfile(path)
except Exception as e:
    print(f"An error occurred: {e}")


# Open the saved workbook using the default program (Excel)
output_path = os.path.join(download_directory, f"Turmas_Unidas_{ano}_{materia}.xlsx")
try:
    subprocess.Popen([output_path], shell=True)
except Exception as e:
    print(f"An error occurred: {e}")

sleep(3)
try:
    os.startfile(new_path)
except Exception as e:
    print(f"An error occurred: {e}")