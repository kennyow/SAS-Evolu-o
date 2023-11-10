import openpyxl
from openpyxl.styles import NamedStyle
import os
import xlwings as xw
import openpyxl
import subprocess
from time import sleep

# Excel path
arquivo_excel = "SAS_EVOLUCAO_2023.xlsx"

# Load the Excel file
wb1 = openpyxl.load_workbook(arquivo_excel)
sheet2 = wb1.active  # Active sheet

# Save all data for each line
dados_main = {}

# # Iterate through all lines
for linha_numero, column in enumerate(sheet2.iter_rows(min_row=2, values_only=True), start=2):
    matricula = column[0]  # Column ⁡⁣⁣⁢A
    nome = column[1]  # Column ⁡⁣⁣B⁡

    # Store the values in the dictionary
    dados_main[linha_numero] = [matricula, nome]

#*Imprimir dados do dicionário
#print(dados_main)
# Caminho do arquivo

file_path = r'C:\Users\citee\Downloads\Simulado SAS Enem 2023 - Edição 5 - Pré-Universitário - Desempenho por Aluno.xlsx'

# Abra o arquivo Excel
workbook = openpyxl.load_workbook(file_path)

# Extrair o nome do arquivo sem extensão
file_name = os.path.splitext(os.path.basename(file_path))[0]

# Dividir o nome do arquivo em partes com base no caractere '-'
parts = file_name.split(' - ')
#*Imprimir a lista das partes do path para renomear o arquivo
#print(parts)


# Selecione a planilha em que você deseja fazer a conversão
sheet = workbook['Alunos']

#Inserindo nova coluna A
sheet.insert_cols(1)

# Inserir o texto "Matrícula" na célula A1
sheet['A1'] = 'Matrícula'


# Iterate through the values of the dictionary
for value in dados_main.values():
    matricula1 = value[0]  # Email from dictionary
    nome1 = value[1]  # Value to insert in column A


    # Iterate through the values in column C of the Excel sheet
    for row_number, row in enumerate(sheet.iter_rows(min_row=0, min_col=2, max_col=2, values_only=True), start=1):
        cell_value = row[0]

        # Compare email_from_dict with cell_value
        if str(nome1) == str(cell_value):
            dictionary_value = value[1]
            sheet.cell(row=row_number, column=1, value=matricula1)

# Defina um estilo personalizado para o novo formato
new_style = NamedStyle(name='new_style')
new_style.number_format = '0.00'

# Itere pelas células da coluna D e faça a conversão
for row in sheet.iter_rows(min_row=2, min_col=5, max_col=8):
    for cell in row:
        # Obtém o valor atual da célula
        cell_value = (cell.value)
        
         # Remove os caracteres não numéricos, como pontos e vírgulas
        cell_value = cell_value.replace('.', '').replace(',', '')
        
        # Converte o valor para float e divide por 100000 para obter o novo formato
        cell_value = float(cell_value) / 10000000
        
        # Define o novo valor formatado e o estilo
        cell.value = cell_value
        cell.style = new_style

# Exclua as colunas C, D, L, M e N
columns_to_delete = [3, 4, 12, 13, 14]  # Lista com os números das colunas a serem excluídas

# Ordene a lista em ordem decrescente para evitar problemas de índice ao excluir colunas
columns_to_delete.sort(reverse=True)

for column_index in columns_to_delete:
    sheet.delete_cols(column_index, 1)

# Salve o arquivo com as alteraçõeso
workbook.save(rf'C:\Users\citee\Downloads\{parts[0]+" " + parts[1]}.xlsx')