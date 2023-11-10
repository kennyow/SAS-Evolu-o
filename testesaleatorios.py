import openpyxl
from openpyxl.utils import get_column_letter

###################################################################################################

# Excel path
arquivo_excel = "SAS_EVOLUCAO_2023.xlsx"

# Load the Excel file
wb1 = openpyxl.load_workbook(arquivo_excel)
sheet2 = wb1.active  # Supondo que você está trabalhando na planilha ativa

# Save all data for each line
dados_main = {}

# # Iterate through all lines
for linha_numero, column in enumerate(sheet2.iter_rows(min_row=2, values_only=True), start=2):
    email_column_b = column[1]  # Column b
    matricula_column_d = column[3]  # Column d

    # Armazenar os valores no dicionário
    dados_main[linha_numero] = [email_column_b, matricula_column_d]

# Closing the excel file
wb1.close()


#############################################################################################


# Criar uma cópia do arquivo original
wb2 = openpyxl.load_workbook(r'C:\Users\citee\Downloads\Turmas_Unidas_6_matematica.xlsx')
sheet = wb2.active  #Active sheet


# Percorra as células da coluna C e exclua o conteúdo
for row in sheet.iter_rows(min_row=1, min_col=3, max_col=3):
    for cell in row:
        cell.value = None


# Iterate through the values of the dictionary
for value in dados_main.values():
    email_from_dict = value[0]  # Email from dictionary
    dictionary_value = value[1]  # Value to insert in column A


    # Iterate through the values in column C of the Excel sheet
    for row_number, row in enumerate(sheet.iter_rows(min_row=0, min_col=2, max_col=2, values_only=True), start=1):
        cell_value = row[0]

        # Compare email_from_dict with cell_value
        if str(email_from_dict) == str(cell_value):
            dictionary_value = value[1]
            sheet.cell(row=row_number, column=3, value=dictionary_value)

# Save the new Excel file

wb2.save(r'C:\Users\citee\Downloads\Turmas_Unidas_6_matematica.xlsx')
wb2.close()